"""Excel exports for report views."""
from __future__ import annotations

import re
from copy import copy
from io import BytesIO

from . import config, db

TEMPLATE = config.ASSETS_DIR / "shablon.xlsx"
LEGACY_TEMPLATE = config.BASE_DIR / "shablon.xlsx"
OBSHIY_TEMPLATE = config.ASSETS_DIR / "obshiy_ves_shablon.xlsx"
LEGACY_OBSHIY_TEMPLATE = config.BASE_DIR / "obshiy_ves_shablon.xlsx"
UMUMIY_TEMPLATE = config.ASSETS_DIR / "umumiy_hisobot_shabloni.xlsx"
LEGACY_UMUMIY_TEMPLATE = config.BASE_DIR / "umumiy_hisobot_shabloni.xlsx"
OBSHIY_ACTION_ORDER = ("top", "topchiqgan", "bizda", "chiqgan")


def _safe_sheet_name(name: str) -> str:
    name = re.sub(r"[\[\]:*?/\\]", " ", (name or "Hisobot")).strip()
    return (name[:31] or "Hisobot")


def _safe_filename(name: str) -> str:
    name = re.sub(r'[<>:"/\\|?*\x00-\x1f]', " ", (name or "Hisobot")).strip()
    name = re.sub(r"\s+", " ", name)
    return f"{name or 'Hisobot'} KARGOLARGA TARQATISH.xlsx"


def _safe_obshiy_filename(name: str) -> str:
    name = re.sub(r'[<>:"/\\|?*\x00-\x1f]', " ", (name or "Hisobot")).strip()
    name = re.sub(r"\s+", " ", name)
    return f"{name or 'Hisobot'} OBSHIY VES.xlsx"


def _safe_umumiy_filename(name: str) -> str:
    name = re.sub(r'[<>:"/\\|?*\x00-\x1f]', " ", (name or "Hisobot")).strip()
    name = re.sub(r"\s+", " ", name)
    return f"{name or 'Hisobot'} UMUMIY HISOBOT.xlsx"


def _num(v) -> float:
    try:
        return round(float(v or 0), 4)
    except (TypeError, ValueError):
        return 0.0


def _formula_num(v) -> str:
    return f"{_num(v):.10g}"


def _entry_expr_and_value(entry: dict) -> tuple[str, float]:
    weight = _num(entry.get("weight"))
    coef = _num(entry.get("coefficient"))
    net = _num(entry.get("net"))
    if coef > 0:
        if not net:
            net = round(weight - coef, 4)
        return f"{_formula_num(weight)}-{_formula_num(coef)}", net
    return _formula_num(net), net


def _slot_value(slot: dict) -> float:
    return round(float(slot.get("value") or 0) + sum(float(x) for x in slot.get("deltas", [])), 4)


def _slot_cell_value(slot: dict):
    expr = str(slot["expr"])
    deltas = list(slot.get("deltas") or [])
    if not deltas and not slot.get("force_formula"):
        return _num(slot.get("value"))
    parts = [expr]
    for delta in deltas:
        n = _formula_num(abs(delta))
        parts.append(("-" if delta < 0 else "+") + n)
    return "=" + "".join(parts)


def _append_reys(slots: dict[str, list[dict]], entry: dict) -> None:
    tovar_turi = str(entry.get("tovar_turi") or "").strip()
    if not tovar_turi:
        return
    expr, value = _entry_expr_and_value(entry)
    slots.setdefault(tovar_turi, []).append({
        "expr": expr,
        "value": value,
        "deltas": [],
        "force_formula": str(expr) != _formula_num(value),
    })


def _apply_adjust(slots: dict[str, list[dict]], entry: dict) -> None:
    from_type = str(entry.get("from_type") or "").strip()
    to_type = str(entry.get("to_type") or "").strip()
    weight = _num(entry.get("weight"))
    if not from_type or not to_type or weight <= 0:
        return

    slots.setdefault(to_type, []).append({
        "expr": _formula_num(weight),
        "value": weight,
        "deltas": [],
        "force_formula": False,
    })

    remaining = weight
    source = slots.setdefault(from_type, [])
    for slot in reversed(source):
        if remaining <= 0:
            break
        available = max(0.0, _slot_value(slot))
        if available <= 0:
            continue
        take = min(available, remaining)
        slot.setdefault("deltas", []).append(-take)
        slot["force_formula"] = True
        remaining = round(remaining - take, 4)
    if remaining > 0:
        source.append({
            "expr": "0",
            "value": 0.0,
            "deltas": [-remaining],
            "force_formula": True,
        })


def _obshiy_code(entry: dict) -> str:
    return (str(entry.get("tovar_turi") or "").strip() or "Kodsiz")


def _sum_by_code(entries: list[dict]) -> tuple[dict[str, float], list[str]]:
    totals: dict[str, float] = {}
    order: list[str] = []
    for entry in entries:
        code = _obshiy_code(entry)
        if code not in totals:
            totals[code] = 0.0
            order.append(code)
        totals[code] = round(totals[code] + _num(entry.get("weight")), 4)
    return totals, order


def _obshiy_value(entry: dict) -> float:
    net = entry.get("net")
    if net is not None:
        return _num(net)
    return round(_num(entry.get("weight")) - _num(entry.get("coefficient")), 4)


def _sum_obshiy_values_by_code(entries: list[dict]) -> tuple[dict[str, float], list[str]]:
    totals: dict[str, float] = {}
    order: list[str] = []
    for entry in entries:
        code = _obshiy_code(entry)
        if code not in totals:
            totals[code] = 0.0
            order.append(code)
        totals[code] = round(totals[code] + _obshiy_value(entry), 4)
    return totals, order


def _ordered_codes(*orders: list[str]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for order in orders:
        for code in order:
            if code not in seen:
                seen.add(code)
                out.append(code)
    return out


def _obshiy_rows(base: dict[str, float], plus: dict[str, float], minus: dict[str, float],
                 order: list[str]) -> list[tuple[str, float, float]]:
    rows = []
    for code in order:
        base_value = _num(base.get(code, 0))
        transfer = round(_num(plus.get(code, 0)) - _num(minus.get(code, 0)), 4)
        if base_value or transfer:
            rows.append((code, base_value, transfer))
    return rows


def _write_obshiy_sheet(ws, rows: list[tuple[str, float, float]], transfer_header: dict[str, str]) -> None:
    from openpyxl.styles import Alignment

    ws.title = transfer_header["sheet"]
    headers = ["Karobka kodi", "karobka kg", transfer_header["column"], "jami"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col)
        cell.value = header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    clear_until = max(ws.max_row, len(rows) + 1)
    for r in range(2, clear_until + 1):
        for c in range(1, 5):
            ws.cell(r, c).value = None

    for idx, (code, base_value, transfer) in enumerate(rows, start=2):
        ws.cell(idx, 1).value = code
        ws.cell(idx, 2).value = base_value
        ws.cell(idx, 3).value = transfer
        ws.cell(idx, 4).value = None
        for col in (2, 3):
            ws.cell(idx, col).number_format = "0.00"

    total_row = 2
    last_row = max(total_row, len(rows) + 1)
    total_cell = ws.cell(total_row, 4)
    total_cell.value = f"=SUM(B{total_row}:C{last_row})" if rows else 0
    total_cell.number_format = "0.00"


def build_obshiy_excel(report_id: int) -> tuple[bytes, str]:
    from openpyxl import Workbook, load_workbook

    report_name = db.report_name(report_id) or "Hisobot"
    entries_by_action = {
        action: list(reversed(db.list_entries(report_id, action, limit=2000)))
        for action in OBSHIY_ACTION_ORDER
    }
    top, top_order = _sum_obshiy_values_by_code(entries_by_action["top"])
    topchiqgan, topchiqgan_order = _sum_obshiy_values_by_code(entries_by_action["topchiqgan"])
    bizda, bizda_order = _sum_obshiy_values_by_code(entries_by_action["bizda"])
    chiqgan, chiqgan_order = _sum_obshiy_values_by_code(entries_by_action["chiqgan"])

    top_rows = _obshiy_rows(
        top,
        plus=chiqgan,
        minus=topchiqgan,
        order=_ordered_codes(top_order, chiqgan_order, topchiqgan_order),
    )
    bizda_rows = _obshiy_rows(
        bizda,
        plus=topchiqgan,
        minus=chiqgan,
        order=_ordered_codes(bizda_order, topchiqgan_order, chiqgan_order),
    )

    template = OBSHIY_TEMPLATE if OBSHIY_TEMPLATE.exists() else LEGACY_OBSHIY_TEMPLATE
    if template.exists():
        wb = load_workbook(template)
    else:
        wb = Workbook()
        wb.active.title = "top"
        wb.create_sheet("bizda qoladigan")

    while len(wb.worksheets) < 2:
        wb.create_sheet("bizda qoladigan" if len(wb.worksheets) == 1 else f"Hisobot {len(wb.worksheets) + 1}")

    _write_obshiy_sheet(
        wb.worksheets[0],
        top_rows,
        {"sheet": "top", "column": "bizdan chiqgan"},
    )
    _write_obshiy_sheet(
        wb.worksheets[1],
        bizda_rows,
        {"sheet": "bizda qoladigan", "column": "topdan chiqgan"},
    )
    for ws in wb.worksheets[2:]:
        wb.remove(ws)

    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    out = BytesIO()
    wb.save(out)
    return out.getvalue(), _safe_obshiy_filename(report_name)


def _summary_type_key(tovar_turi: str) -> str:
    key = str(tovar_turi or "").strip().lower()
    if key == "one":
        return "oneway"
    if key == "uztez":
        return "uzt"
    if key.startswith("x"):
        return "xabib"
    return key


def _inventory_for_summary(report_id: int) -> dict[str, float]:
    inv: dict[str, float] = {}
    for tovar_turi, value in db.get_inventory(report_id).items():
        key = _summary_type_key(tovar_turi)
        if not key:
            continue
        inv[key] = round(inv.get(key, 0) + _num(value), 4)
    return inv


def _copy_row_style(ws, source_row: int, target_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        dst.number_format = src.number_format
        dst.alignment = copy(src.alignment)
        dst.fill = copy(src.fill)
        dst.font = copy(src.font)
        dst.border = copy(src.border)


def build_umumiy_excel(report_id: int) -> tuple[bytes, str]:
    from datetime import date
    from openpyxl import Workbook, load_workbook

    report_name = db.report_name(report_id) or "Hisobot"
    obshiy_entries = {
        action: list(reversed(db.list_entries(report_id, action, limit=2000)))
        for action in OBSHIY_ACTION_ORDER
    }
    topchiqgan, topchiqgan_order = _sum_obshiy_values_by_code(obshiy_entries["topchiqgan"])
    bizda, bizda_order = _sum_obshiy_values_by_code(obshiy_entries["bizda"])
    chiqgan, chiqgan_order = _sum_obshiy_values_by_code(obshiy_entries["chiqgan"])
    bizda_rows = _obshiy_rows(
        bizda,
        plus=topchiqgan,
        minus=chiqgan,
        order=_ordered_codes(bizda_order, topchiqgan_order, chiqgan_order),
    )

    bizda_total = round(sum(round(base + transfer, 4) for _, base, transfer in bizda_rows), 4)
    box_weight_total = round(sum(
        _num(e.get("coefficient"))
        for entries in obshiy_entries.values()
        for e in entries
    ), 4)
    inv = _inventory_for_summary(report_id)
    top_inventory = _num(inv.get("top", 0))
    inv["top"] = 0.0

    template = UMUMIY_TEMPLATE if UMUMIY_TEMPLATE.exists() else LEGACY_UMUMIY_TEMPLATE
    if template.exists():
        wb = load_workbook(template)
    else:
        wb = Workbook()
    ws = wb.active
    ws.title = _safe_sheet_name(report_name)

    max_clear_row = max(ws.max_row, 129)
    for r in range(2, max_clear_row + 1):
        ws.cell(r, 1).value = None

    a_row = 2
    clean_total = round(bizda_total - top_inventory, 4)
    ws.cell(a_row, 1).value = clean_total
    ws.cell(a_row, 1).number_format = "0.00"

    ws["E2"] = "To'lashi kerak bo'lgan summa:"
    ws["F2"] = date.today().strftime("%d.%m.%Y")
    ws["F3"] = report_name
    ws["C2"] = "=SUM(A:A)"
    ws["C3"] = box_weight_total
    ws["C3"].number_format = "0.00"

    label_rows: dict[str, int] = {}
    for row in range(4, ws.max_row + 1):
        raw_label = str(ws.cell(row, 2).value or "").strip().lower()
        label = _summary_type_key(raw_label)
        if label:
            if raw_label == "one":
                ws.cell(row, 2).value = "oneway"
            label_rows[label] = row

    represented = set(label_rows)
    custom_types = [
        t for t, value in inv.items()
        if value and t not in represented and t not in {"mandarin", "uztez"}
    ]
    next_row = max([r for r in label_rows.values()] + [15]) + 1
    for tovar_turi in custom_types:
        _copy_row_style(ws, 15, next_row, 8)
        ws.cell(next_row, 2).value = tovar_turi
        label_rows[tovar_turi] = next_row
        next_row += 1

    distributed_labels = {"akb", "jet", "xabib", "navo", "jon", "oneway", "redwing", "uzt"}
    for label, row in label_rows.items():
        if label == "karobka":
            continue
        if label == "mandarin":
            continue
        ws.cell(row, 3).value = inv.get(label, 0)
        ws.cell(row, 3).number_format = "0.00"

    distributable_rows = [
        row for label, row in label_rows.items()
        if label in distributed_labels
    ]
    non_distributed_rows = [
        row for label, row in label_rows.items()
        if label not in distributed_labels and label not in {"karobka", "mandarin"}
    ]
    for row in non_distributed_rows:
        ws.cell(row, 4).value = None
        ws.cell(row, 5).value = f"=C{row}+D{row}"

    mandarin_row = label_rows.get("mandarin", 4)
    non_distributed_refs = [f"C{row}" for row in non_distributed_rows]
    ws["G3"] = f"=C2" + ("-" + "-".join(non_distributed_refs) if non_distributed_refs else "") + "-C3"
    ws["H3"] = "=IF(G3=0,0,C3/G3)"
    ws.cell(mandarin_row, 3).value = (
        "=G3" + ("-" + "-".join(f"C{row}" for row in distributable_rows) if distributable_rows else "")
    )
    ws.cell(mandarin_row, 4).value = f"=C{mandarin_row}*$H$3"
    ws.cell(mandarin_row, 5).value = f"=C{mandarin_row}+D{mandarin_row}"

    for row in distributable_rows:
        ws.cell(row, 4).value = f"=C{row}*$H$3"
        ws.cell(row, 5).value = f"=C{row}+D{row}"

    for row in set([3, mandarin_row] + distributable_rows + non_distributed_rows):
        for col in (3, 4, 5, 7, 8):
            ws.cell(row, col).number_format = "0.00"

    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    out = BytesIO()
    wb.save(out)
    return out.getvalue(), _safe_umumiy_filename(report_name)


def build_kargo_excel(report_id: int) -> tuple[bytes, str]:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    report_name = db.report_name(report_id) or "Hisobot"
    entries = list(reversed(db.list_entries(report_id, "reys", limit=2000)))
    adjusts = list(reversed(db.list_entries(report_id, "adjust", limit=2000)))

    slots: dict[str, list[dict]] = {}
    custom_types: list[str] = []
    default_set = set(db.DEFAULT_TYPES)
    for entry in entries:
        tovar_turi = str(entry.get("tovar_turi") or "").strip()
        if not tovar_turi:
            continue
        _append_reys(slots, entry)
        if tovar_turi not in default_set and tovar_turi not in custom_types:
            custom_types.append(tovar_turi)
    for entry in adjusts:
        _apply_adjust(slots, entry)
        for tovar_turi in (str(entry.get("from_type") or "").strip(), str(entry.get("to_type") or "").strip()):
            if tovar_turi and tovar_turi not in default_set and tovar_turi not in custom_types:
                custom_types.append(tovar_turi)

    template = TEMPLATE if TEMPLATE.exists() else LEGACY_TEMPLATE
    if template.exists():
        wb = load_workbook(template)
    else:
        wb = Workbook()
        ws0 = wb.active
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for r, tovar_turi in enumerate(db.DEFAULT_TYPES, start=1):
            ws0.cell(r, 1).value = tovar_turi
            ws0.cell(r, 1).fill = PatternFill("solid", fgColor="FFFF00")
            ws0.cell(r, 1).font = Font(bold=True)
            ws0.cell(r, 1).alignment = Alignment(horizontal="center", vertical="center")
            ws0.cell(r, 1).border = border
            ws0.cell(r, 2).number_format = "0.00"
            ws0.cell(r, 2).border = border
        ws0.column_dimensions["A"].width = 18
        for col in range(2, 30):
            ws0.column_dimensions[get_column_letter(col)].width = 12

    ws = wb.active
    ws.title = _safe_sheet_name(report_name)

    label_style = copy(ws["A1"]._style)
    custom_label_style = copy(ws["A13"]._style if ws["A13"].has_style else ws["A1"]._style)
    value_style = copy(ws["B1"]._style)
    blank_style = copy(ws["K1"]._style)
    label_fill = copy(ws["A1"].fill)
    value_num_format = ws["B1"].number_format

    types = list(db.DEFAULT_TYPES)
    data_rows: list[tuple[int, str, bool]] = []
    row = 1
    for tovar_turi in types:
        data_rows.append((row, tovar_turi, False))
        row += 1
    if custom_types:
        row += 1
        for tovar_turi in custom_types:
            data_rows.append((row, tovar_turi, True))
            row += 1

    max_entries = max([len(slots.get(t, [])) for _, t, _ in data_rows] + [1])
    max_col = max(2, 1 + max_entries)
    total_start = (data_rows[-1][0] if data_rows else 1) + 6
    total_end = total_start + len(data_rows) - 1
    clear_rows = max(ws.max_row, total_end)
    clear_cols = max(ws.max_column, max_col)

    for clear_row in range(1, clear_rows + 1):
        for clear_col in range(1, clear_cols + 1):
            cell = ws.cell(clear_row, clear_col)
            cell.value = None
            cell._style = copy(blank_style)

    data_row_by_type: dict[str, int] = {}
    for data_row, tovar_turi, is_custom in data_rows:
        data_row_by_type[tovar_turi] = data_row
        cell = ws.cell(data_row, 1)
        cell.value = tovar_turi
        cell._style = copy(custom_label_style if is_custom else label_style)
        cell.fill = copy(label_fill)
        for idx, slot in enumerate(slots.get(tovar_turi, []), start=2):
            val_cell = ws.cell(data_row, idx)
            val_cell.value = _slot_cell_value(slot)
            val_cell._style = copy(value_style)
            val_cell.number_format = value_num_format

    total_row = total_start
    last_col = get_column_letter(max_col)
    for _, tovar_turi, is_custom in data_rows:
        a = ws.cell(total_row, 1)
        b = ws.cell(total_row, 2)
        a.value = tovar_turi
        a._style = copy(custom_label_style if is_custom else label_style)
        a.fill = copy(label_fill)
        b.value = f"=SUM(B{data_row_by_type[tovar_turi]}:{last_col}{data_row_by_type[tovar_turi]})"
        b._style = copy(value_style)
        b.number_format = value_num_format
        total_row += 1

    ws.sheet_view.showGridLines = True
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    out = BytesIO()
    wb.save(out)
    return out.getvalue(), _safe_filename(report_name)
